import xlrd
import openpyxl
from pathlib import Path
from colour import Color
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.units import cm_to_EMU, pixels_to_EMU
from xlrd import XLRDError
from utils.files import open_file
from utils.mylogging import error
from utils.progress_count import ProgressCount


def get_excel_data_given_format(fname: str) -> list:
    try:
        wb = xlrd.open_workbook(fname, formatting_info=True)
        sh = wb.sheet_by_index(0)
        rows = []
        for rowx in range(sh.nrows):
            row = []
            for colx in range(sh.ncols):
                cell = sh.cell(rowx, colx)
                xf = wb.xf_list[cell.xf_index]

                format = wb.format_map[xf.format_key]
                format_str = format.format_str

                if cell.ctype == xlrd.XL_CELL_NUMBER and all(x == '0' for x in format_str):
                    row.append('%0*d' % (len(format_str), int(cell.value)))
                else:
                    row.append(str(cell.value))

            rows.append(row)
        return rows
    except XLRDError:
        error(f'Unsupported format, or corrupt file: {fname}')
        exit()


def get_excel_data_indent_levels(fname: str) -> list:
    try:
        wb = xlrd.open_workbook(fname, formatting_info=True)
        sh = wb.sheet_by_index(0)
        rows = []
        for rowx in range(sh.nrows):
            row = []
            for colx in range(sh.ncols):
                cell = sh.cell(rowx, colx)
                xf = wb.xf_list[cell.xf_index]
                row.append(xf.alignment.indent_level)

            rows.append(row)
        return rows
    except XLRDError:
        error(f'Unsupported format, or corrupt file: {fname}')
        exit()


def get_excel_data_with_headers_given_format(fname: str) -> (list, list):
    try:
        wb = xlrd.open_workbook(fname, formatting_info=True)
        sh = wb.sheet_by_index(0)

        headers_dict = {}
        for colx in range(sh.ncols):
            heading = str(sh.cell(0, colx).value)
            if len(heading) > 0:
                headers_dict[colx] = heading

        rows = []

        for rowx in range(sh.nrows)[1:]:
            row = {}

            for colx in headers_dict.keys():
                cell = sh.cell(rowx, colx)
                xf = wb.xf_list[cell.xf_index]

                format = wb.format_map[xf.format_key]
                format_str = format.format_str

                if cell.ctype == xlrd.XL_CELL_NUMBER and all(x == '0' for x in format_str):
                    val = '%0*d' % (len(format_str), int(cell.value))
                else:
                    val = str(cell.value)

                row[headers_dict[colx]] = val

            rows.append(row)

        headers = []
        for h in headers_dict.values():
            headers.append(h)

        return headers, rows
    except XLRDError:
        error(f'Unsupported format, or corrupt file: {fname}')
        exit()


def get_xlsx_data(fname, *, data_only=True):
    wb = openpyxl.load_workbook(fname, data_only=data_only)
    ws = wb.active
    rows = []
    for row in ws.iter_rows():
        row_data = []
        for cell in row:
            row_data.append(cell.value)
        rows.append(row_data)
    return rows


def width_auto(ws, column, add=0):
    col = get_column_letter(column) if type(column) is int else column
    max_length = 0
    for cell in ws[f'{col}:{col}']:
        # try:  # Necessary to avoid error on empty cells
        str_value = str(cell.value)
        len_cell_value = len(str_value)
        if len_cell_value > max_length:
            max_length = len_cell_value
        # except:
        #     pass
    # adjusted_width = (max_length + 2) * 1.2
    adjusted_width = max_length + 1 + add
    ws.column_dimensions[col].width = adjusted_width


def row_font_alignment(ws, row, font, alignment):
    for cell in ws[f'{row}:{row}']:
        cell.font = font
        cell.alignment = alignment


def header_row_font(ws, row):
    row_font_alignment(ws, row, Font(bold=True), Alignment(horizontal='center', vertical='center', wrap_text=True))


def column_bold(ws, column):
    col = get_column_letter(column) if type(column) is int else column
    for cell in ws[f'{col}:{col}']:
        cell.font = Font(bold=True)


def column_color(ws, column, color):
    col = get_column_letter(column) if type(column) is int else column
    col_hex = Color(color).hex_l[1:]
    for cell in ws[f'{col}:{col}']:
        cell.font = Font(color=col_hex)


def column_alignment(ws, column, horizontal='center', vertical='center', wrap_text=True):
    col = get_column_letter(column) if type(column) is int else column
    for cell in ws[f'{col}:{col}']:
        cell.alignment = Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text)


def column_background(ws, column, color):
    col = get_column_letter(column) if type(column) is int else column
    col_hex = Color(color).hex_l[1:]
    for cell in ws[f'{col}:{col}']:
        cell.fill = PatternFill(start_color=col_hex, end_color=col_hex, fill_type="solid")


def cell_background(cell, color):
    col_hex = Color(color).hex_l[1:]
    cell.fill = PatternFill(start_color=col_hex, end_color=col_hex, fill_type='solid')


def cell_color(cell, color):
    col_hex = Color(color).hex_l[1:]
    cell.font = Font(color=col_hex)


def columns_width(ws, col_start, col_finish, width):
    for c in range(col_start, col_finish + 1):
        col = get_column_letter(c)
        ws.column_dimensions[col].width = width


def create_excel_with_data(filename: str, data: list, startfile: bool = False, template=None,
                           data_sh_name: str = None) -> None:
    if template:
        xlsx = load_workbook(template)
        xlsx.__class__ = Xlsx
        xlsx.filename = filename
        if data_sh_name:
            xlsx.change_sheet(data_sh_name)
    else:
        xlsx = Xlsx(filename, overwrite=True)

    xlsx.set_data(data).save()
    if startfile:
        xlsx.open_file()


class Xlsx(Workbook):
    def __init__(self, filename, overwrite=False, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.filename = filename

        if Path(self.filename).exists() and not overwrite:
            error(f'File exists: {filename}')
            exit()

        self.data = None
        self.ws = self.active

    def save(self, *args, **kwargs):
        super(Xlsx, self).save(self.filename)
        return self

    def open_file(self):
        open_file(self.filename)

    def save_and_open(self):
        self.save()
        self.open_file()

    def set_data(self, data: list):
        self.data = data
        pr = ProgressCount(len(data))
        for row in data:
            pr.next()
            self.ws.append(row)
        return self

    # def openfolder(self):
    #     if not is_mac():
    #         os.startfile(directory_path(self.filename))

    def convert_cols_arg_to_list_cols(self, cols: any) -> list:
        if not cols:
            cols = f'A:{get_column_letter(self.ws.max_column)}'
        else:
            cols = get_column_letter(cols) if type(cols) is int else cols
        list_cols = []
        spl1 = cols.split(',')
        for s in spl1:
            spl2 = s.split(':')
            start_index = column_index_from_string(spl2[0].strip())
            end_index = (column_index_from_string(spl2[1].strip()) + 1) if len(spl2) == 2 else (start_index + 1)
            for i in range(start_index, end_index):
                list_cols.append(get_column_letter(i))
        return list_cols

    def width_auto(self, cols: any = None):
        """
        If cols is None then set width auto everything [A:{ws.max_column}]
        Examples
        cols = 'A:C,E:F'
        cols = 5
        """
        list_cols = self.convert_cols_arg_to_list_cols(cols)
        for col in list_cols:
            max_length = 0
            for cell in self.ws[f'{col}:{col}']:
                # try:  # Necessary to avoid error on empty cells
                str_value = str(cell.value)
                len_cell_value = len(str_value)
                if len_cell_value > max_length:
                    max_length = len_cell_value
                # except:
                #     pass
            adjusted_width = (max_length + 2) * 1.1
            self.ws.column_dimensions[col].width = adjusted_width
        return self

    def column_width(self, col: int, width: float):
        column_letter = get_column_letter(col)
        self.ws.column_dimensions[column_letter].width = width
        return self

    def columns_width(self, cols: any, width: float = 8):
        list_cols = self.convert_cols_arg_to_list_cols(cols)
        for col in list_cols:
            self.ws.column_dimensions[col].width = width
        return self

    def row_height(self, row, height):
        self.ws.row_dimensions[row].height = height
        return self

    def cell_bold(self, row, col):
        self.ws.cell(row=row, column=col).font = Font(bold=True)

    def cell_alignment(self, row, col, horizontal: str = 'center', vertical: str = 'center', wrap_text: bool = True):
        self.ws.cell(row=row, column=col).alignment = \
            Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text)
        return self

    def columns_alignment(self, cols: any, horizontal: str = 'center', vertical: str = 'center',
                          wrap_text: bool = True):
        list_cols = self.convert_cols_arg_to_list_cols(cols)
        for col in list_cols:
            for cell in self.ws[f'{col}:{col}']:
                cell.alignment = Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text)
        return self

    def zoom_scale(self, zoom_scale: int = 85):
        self.ws.sheet_view.zoomScale = zoom_scale
        return self

    def create_sheet(self, title=None, index=None):
        self.ws = super(Xlsx, self).create_sheet(title, index)
        return self

    def change_sheet(self, sheet_name):
        self.ws = self[sheet_name]
        return self

    def last_row_font_alignment(self, font: Font = None, alignment: Alignment = None):
        for cell in self.ws[f'{self.ws.max_row}:{self.ws.max_row}']:
            if font:
                cell.font = font
            if alignment:
                cell.alignment = alignment
        return self

    def cell_background(self, row, col, color):
        col_hex = Color(color).hex_l[1:]
        self.ws.cell(row=row, column=col).fill = PatternFill(start_color=col_hex, end_color=col_hex, fill_type='solid')

    def last_row_background_color(self, color_hex: str):
        my_fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
        for cell in self.ws[f'{self.ws.max_row}:{self.ws.max_row}']:
            cell.fill = my_fill
        return self

    def row_font_alignment(self, row, font, alignment):
        for cell in self.ws[f'{row}:{row}']:
            cell.font = font
            cell.alignment = alignment
        return self

    def header_row_font(self, row):
        row_font_alignment(self.ws, row,
                           Font(bold=True),
                           Alignment(horizontal='center', vertical='center', wrap_text=True))
        return self

    @staticmethod
    def load_workbook(filename):
        wb = load_workbook(filename)
        wb.__class__ = Xlsx
        wb.ws = wb.active
        wb.filename = filename
        return wb

    def col_names(self):
        names = []
        for col in range(1, self.ws.max_column + 1):
            names.append(self.ws.cell(row=1, column=col).value)
        if names == [None]:
            names = []
        return names

    def col_num_by_col_name(self, col_name):
        col_names = self.col_names()
        if col_name in col_names:
            col_num = col_names.index(col_name) + 1
        else:
            col_num = len(col_names) + 1
            self.ws.cell(row=1, column=col_num).value = col_name
        return col_num

    def to_col_with_name(self, col_name, value, row):
        col_num = self.col_num_by_col_name(col_name)
        self.ws.cell(row=row, column=col_num).value = value

    # Image

    def img_to_cell(self, row, col, img_path, coloffset=0.1, rowoffset=0.1):
        img = openpyxl.drawing.image.Image(img_path)
        img.anchor = f'{get_column_letter(col)}{row}'

        p2e = pixels_to_EMU
        h, w = img.height, img.width
        size = XDRPositiveSize2D(p2e(w), p2e(h))

        c2e = cm_to_EMU
        cellh = lambda x: c2e((x * 49.77) / 99)
        cellw = lambda x: c2e((x * (18.65 - 1.71)) / 10)
        coloffset = cellw(coloffset)
        rowoffset = cellh(rowoffset)
        marker = AnchorMarker(col=col - 1, colOff=coloffset, row=row - 1, rowOff=rowoffset)
        img.anchor = OneCellAnchor(_from=marker, ext=size)

        self.ws.add_image(img)
        del img

    def img_to_col_with_name(self, col_name, img_path, row):
        col_num = self.col_num_by_col_name(col_name)
        self.img_to_cell(row, col_num, img_path)


class XlsxFromTemplate(Xlsx):
    def __init__(self):
        super(XlsxFromTemplate, self).__init__()


def load_workbook_and_width_auto(filepath, then_open=False):
    xlsx = Xlsx.load_workbook(filepath)
    assert isinstance(xlsx, Xlsx)
    xlsx.width_auto().save()
    if then_open:
        xlsx.save_and_open()
